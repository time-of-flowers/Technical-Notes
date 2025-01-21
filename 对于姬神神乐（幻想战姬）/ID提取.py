import os
import openpyxl

def extract_numbers_from_sheet(sheet):
    numbers = {"1": [], "2": []}  # 用于存储以1和2开头的数字
    for row in sheet.iter_rows(min_col=1, max_col=1):  # 只读取A列
        for cell in row:
            # 打印出A列中的值，检查它们的格式
            print(f"读取的单元格值: {cell.value} (类型: {type(cell.value)})")
            
            # 如果是浮动点类型，处理为整数
            if isinstance(cell.value, float):
                # 转换为整数
                int_value = int(cell.value)
                # 确保是5位数字并符合以1或2开头的条件
                if len(str(int_value)) == 5 and str(int_value).startswith(("1", "2")):
                    if str(int_value).startswith("1"):
                        numbers["1"].append(int_value)
                    elif str(int_value).startswith("2"):
                        numbers["2"].append(int_value)
    
    # 打印提取的数据
    print(f"提取到的以1开头的数字: {numbers['1']}")
    print(f"提取到的以2开头的数字: {numbers['2']}")
    return numbers

def process_xlsx_files(file1, file2, output_file):
    # 检查文件路径
    print("file1路径:", file1)
    print("file2路径:", file2)
    
    try:
        wb1 = openpyxl.load_workbook(file1)
        wb2 = openpyxl.load_workbook(file2)
    except Exception as e:
        print(f"加载文件时出错: {e}")
        return
    
    numbers = {"1": [], "2": []}  # 用于存储最终的结果

    # 打印工作表名称以确保文件被正确加载
    print(f"文件 {file1} 的工作表: {wb1.sheetnames}")
    print(f"文件 {file2} 的工作表: {wb2.sheetnames}")
    
    # 遍历每个文件中的每一页
    for sheet_name in wb1.sheetnames:
        sheet = wb1[sheet_name]
        extracted = extract_numbers_from_sheet(sheet)
        numbers["1"].extend(extracted["1"])
        numbers["2"].extend(extracted["2"])

    for sheet_name in wb2.sheetnames:
        sheet = wb2[sheet_name]
        extracted = extract_numbers_from_sheet(sheet)
        numbers["1"].extend(extracted["1"])
        numbers["2"].extend(extracted["2"])

    # 写入txt文件
    if numbers["1"] or numbers["2"]:
        with open(output_file, 'w') as f:
            f.write("以1开头的数字:\n")
            for num in numbers["1"]:
                f.write(f"{num}\n")
            f.write("\n以2开头的数字:\n")
            for num in numbers["2"]:
                f.write(f"{num}\n")
        print(f"数据已经成功写入 {output_file}")
    else:
        print("没有找到符合条件的数据")

# 输入文件和输出文件路径
file1 = r"F:\$Users_V\Desktop\姬神能力圖鑑及技能突破.xlsx"
file2 = r"F:\$Users_V\Desktop\物靈能力圖鑑及技能突破.xlsx"
output_file = "output.txt"

process_xlsx_files(file1, file2, output_file)
